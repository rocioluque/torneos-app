from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


CAT_COLORS = {
    "2012": "C8E6C9",  # verde
    "2013": "B3E5FC",  # celeste
    "2014": "FFF9C4",  # amarillo
    "2015": "F8BBD0",  # rosa
}

HEADER_VIERNES = "B3D9FF"
HEADER_SABADO = "FFE0B2"
GRAY_EMPTY = "F5F5F5"
THICK = Side(style="thick", color="000000")
THIN = Side(style="thin", color="CCCCCC")


def _border_thick_bottom():
    return Border(bottom=THICK, top=THIN, left=THIN, right=THIN)


def _border_normal():
    return Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def generar_programacion_excel(resultado: dict, horarios: list[str], horarios_extra: list[str], total_canchas: int) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Programación"

    headers = ["N°", "DÍA", "HORA", "CANCHA", "CATEGORÍA", "EQUIPO 1", "RESULTADO", "EQUIPO 2", "ZONA", "FECHA"]
    col_widths = [5, 10, 8, 8, 10, 22, 12, 22, 12, 8]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="37474F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 20

    all_horarios = horarios + horarios_extra
    row_num = 2
    num = 1

    for day in ("viernes", "sabado"):
        matches = resultado.get(day, [])
        if not matches:
            continue

        day_label = "VIERNES" if day == "viernes" else "SÁBADO"
        day_color = HEADER_VIERNES if day == "viernes" else HEADER_SABADO

        # Encabezado de día
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(headers))
        cell = ws.cell(row=row_num, column=1, value=day_label)
        cell.font = Font(bold=True, size=13)
        cell.fill = PatternFill("solid", fgColor=day_color)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_num].height = 18
        row_num += 1

        # Agrupar por turno
        turnos_usados = sorted(set(m["turno"] for m in matches))

        for turno in turnos_usados:
            hora = all_horarios[turno - 1] if turno <= len(all_horarios) else f"T{turno}"
            turno_matches = {m["cancha"]: m for m in matches if m["turno"] == turno}

            last_row_of_turno = row_num + total_canchas - 1

            for cancha in range(1, total_canchas + 1):
                m = turno_matches.get(cancha)
                is_last = cancha == total_canchas
                border = _border_thick_bottom() if is_last else _border_normal()

                if m:
                    color = CAT_COLORS.get(str(m["cat"]), "FFFFFF")
                    fill = PatternFill("solid", fgColor=color)
                    values = [num, day_label, hora, cancha, m["cat"], m["t1"], "", m["t2"], m.get("group", ""), m.get("fecha", "")]
                    for col, val in enumerate(values, 1):
                        cell = ws.cell(row=row_num, column=col, value=val)
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border
                    num += 1
                else:
                    fill = PatternFill("solid", fgColor=GRAY_EMPTY)
                    values = ["", day_label, hora, cancha, "", "", "", "", "", ""]
                    for col, val in enumerate(values, 1):
                        cell = ws.cell(row=row_num, column=col, value=val)
                        cell.fill = fill
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                        cell.border = border

                row_num += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generar_grilla_categoria_excel(resultado: dict) -> bytes:
    all_matches = resultado.get("viernes", []) + resultado.get("sabado", [])
    categorias = sorted(set(m["cat"] for m in all_matches))

    wb = Workbook()
    wb.remove(wb.active)

    for cat in categorias:
        ws = wb.create_sheet(title=f"Cat {cat}")
        color = CAT_COLORS.get(str(cat), "FFFFFF")

        cat_matches = [m for m in all_matches if m["cat"] == cat]
        equipos = sorted(set(eq for m in cat_matches for eq in (m["t1"], m["t2"])))

        # Detectar cuántos partidos por día
        max_partidos = 3
        headers = ["EQUIPO", "GRUPO"]
        for i in range(1, max_partidos + 1):
            headers += [f"DÍA {i}", f"HORA {i}", f"CANCHA {i}", f"RIVAL {i}"]

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="37474F")
            cell.alignment = Alignment(horizontal="center")

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 10
        for c in range(3, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 12

        row_num = 2
        for eq in equipos:
            eq_matches = sorted(
                [m for m in cat_matches if eq in (m["t1"], m["t2"])],
                key=lambda x: (x["day"] != "viernes", x["turno"])
            )
            group = eq_matches[0].get("group", "") if eq_matches else ""
            row = [eq, group]
            for m in eq_matches[:max_partidos]:
                rival = m["t2"] if m["t1"] == eq else m["t1"]
                row += [
                    "VIERNES" if m["day"] == "viernes" else "SÁBADO",
                    m.get("hora", ""),
                    m.get("cancha", ""),
                    rival,
                ]
            # Rellenar si tiene menos de max_partidos
            while len(row) < 2 + max_partidos * 4:
                row += ["", "", "", ""]

            fill = PatternFill("solid", fgColor=color)
            for col, val in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
            row_num += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
